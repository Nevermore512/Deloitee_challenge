from django.shortcuts import render
from .forms import FileForm
from django.http import HttpResponseRedirect, HttpResponse
import pandas as pd
import numpy as np
import os
from openpyxl import Workbook, load_workbook


def input(request):
    submitted = False
    if request.method == 'POST':
        form = FileForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect('/finance/?submitted=True')
    else:
        form = FileForm()
        print(form.as_table())
        if 'submitted' in request.GET:
            submitted = True
    return render(request, 'finace/input.html',
                  {'form': form, 'submitted': submitted})


def addl(*arg):
    result = []
    for i in arg:
        result.append(i)
    return result


def output(request):
    # 加载文件
    BASE_DIR = r'E:/pycharm_workplace/Deqin'
    bs_name = os.listdir(os.path.join(BASE_DIR, r'upload/balance'))[0]
    bs_dir = os.path.join(os.path.join(BASE_DIR, r'upload/balance')
                          , bs_name)
    # print(bs_name, bs_dir)

    ics_name = os.listdir(os.path.join(BASE_DIR, r'upload/income'))[0]
    ics_dir = os.path.join(os.path.join(BASE_DIR, r'upload/income')
                           , ics_name)

    cf_name = os.listdir(os.path.join(BASE_DIR, r'upload/carsh'))[0]
    cf_dir = os.path.join(os.path.join(BASE_DIR, r'upload/carsh')
                          , cf_name)

    bs = pd.read_csv(bs_dir, encoding='gbk').set_index('报告日期').replace(['--', '\t\t', ' --'], np.nan).astype('float')
    ics = pd.read_csv(ics_dir, encoding='gbk').set_index('报告日期').replace(['--', '\t\t', ' --'], np.nan).astype('float')
    cf = pd.read_csv(cf_dir, encoding='gbk').set_index(' 报告日期').replace(['--', ' ', ' --'], np.nan).astype('float')

    fin_years = ['2018-12-31', '2017-12-31', '2016-12-31']

    asset_sum = bs.loc['资产总计(万元)', fin_years]
    debt_sum = bs.loc['负债合计(万元)', fin_years]
    objects = asset_sum - bs.loc['无形资产(万元)', fin_years]
    d_to_a = debt_sum / asset_sum
    a_to_o = objects / debt_sum
    fin_struct = addl(asset_sum, debt_sum, objects, d_to_a, a_to_o)
    fin_struct = pd.DataFrame(fin_struct, index=['总资产', '总负债', '有形资产', '资产负债率', '财务杠杆比率'])
    # print(fin_struct)

    asset_flu = bs.loc['流动资产合计(万元)', fin_years]
    inventory = bs.loc['存货(万元)', fin_years]
    debt_flu = bs.loc['流动负债合计(万元)', fin_years]
    flu_rate = asset_flu / debt_flu
    aci_rate = (asset_flu - inventory) / flu_rate
    fin_flu = addl(asset_flu, inventory, debt_flu, flu_rate, aci_rate)
    fin_flu = pd.DataFrame(fin_flu, index=['流动资产', '存货', '流动负债', '流动比率', '速动比率'])
    # print(fin_flu)

    inven_rate = ics.loc['营业成本(万元)', fin_years] / inventory
    ys_rate = ics.loc['营业收入(万元)', fin_years] / bs.loc['应收账款(万元)', fin_years]
    yf_rate = ics.loc['营业总成本(万元)', fin_years] / bs.loc['应付账款(万元)', fin_years]
    fin_opr = addl(inven_rate, ys_rate, yf_rate)
    fin_opr = pd.DataFrame(fin_opr, index=['存货周转率', '应收账款周转率', '应付账款周转率'])
    # print(fin_opr)

    fin_cf = cf.loc[[' 经营活动产生的现金流量净额(万元)', ' 投资活动产生的现金流量净额(万元)', ' 筹资活动产生的现金流量净额(万元)'], fin_years]
    c_to_d = cf.loc[' 经营活动产生的现金流量净额(万元)', fin_years] / debt_flu
    fin_cf = fin_cf.append(c_to_d, ignore_index=True)
    fin_cf.index = ['经营活动产⽣的现⾦流量净额', '投资活动产⽣的现⾦流量净额', '筹资活动产生的现金流量净额', '现⾦流动负债⽐率']
    # print(fin_cf)

    income = ics.loc['营业收入(万元)', fin_years]
    net_income = income - ics.loc['营业成本(万元)', fin_years]
    ni_rate = net_income / income
    profit = ics.loc['利润总额(万元)', fin_years]
    pr_rate = profit / income
    net_rate = profit / (asset_sum - debt_sum)
    fin_pro = addl(income, net_income, ni_rate, profit, pr_rate, net_rate)
    fin_pro = pd.DataFrame(fin_pro, index=['营业收入', '毛利', '毛利率', '净利', '净利率', '净资产收益率'])
    # print(fin_pro)

    # 所有者权益(或股东权益)合计(万元)
    s2018 = bs.loc['所有者权益(或股东权益)合计(万元)', '2018-12-31']
    s2017 = bs.loc['所有者权益(或股东权益)合计(万元)', '2017-12-31']
    s2016 = bs.loc['所有者权益(或股东权益)合计(万元)', '2016-12-31']
    i2018 = ics.loc['营业收入(万元)', '2018-12-31']
    i2017 = ics.loc['营业收入(万元)', '2017-12-31']
    i2016 = ics.loc['营业收入(万元)', '2016-12-31']
    p2018 = ics.loc['利润总额(万元)', '2018-12-31']
    p2017 = ics.loc['利润总额(万元)', '2017-12-31']
    p2016 = ics.loc['利润总额(万元)', '2016-12-31']
    asset_cum = pd.Series(addl((s2018 - s2017) / s2017, (s2017 - s2016) / s2016))
    income_in = pd.Series(addl((i2018 - i2017) / i2017, (i2017 - i2016) / i2016))
    profit_in = pd.Series(addl((p2018 - p2017) / p2017, (p2017 - p2016) / p2016))
    fin_inc = pd.DataFrame(addl(asset_cum, income_in, profit_in), index=['资本积累率', '营业收入增长率', '净利润增长率'])
    fin_inc.columns = ['2018', '2017']

    print((fin_inc))

    wb = Workbook()
    struct = wb.create_sheet('资本结构')  # fin_struct
    flu = wb.create_sheet('流动性')  # fin_flu
    opr = wb.create_sheet('经营情况')  # fin_opr
    cf = wb.create_sheet('现金流量')  # fin_cf
    pro = wb.create_sheet('盈利能力')  # fin_pro
    inc = wb.create_sheet('成长性')  # fin_inc
    credit = wb.create_sheet('失信与被执行人')
    wb.save('./upload/report/fin.xlsx')

    with pd.ExcelWriter('./upload/report/fin.xlsx') as writer:
        fin_struct.to_excel(writer, sheet_name='资本结构')
        fin_flu.to_excel(writer, sheet_name='流动性')
        fin_opr.to_excel(writer, sheet_name='经营情况')
        fin_cf.to_excel(writer, sheet_name='现金流量')
        fin_pro.to_excel(writer, sheet_name='盈利能力')
        fin_inc.to_excel(writer, sheet_name='成长性')

        asset = fin_struct.loc['总资产', '2018-12-31']
        debt = fin_struct.loc['总负债', '2018-12-31']
        equilt = asset - debt
        name = '万科集团'
        risk = '低风险'
        re = '中回报'

    wb = load_workbook(filename='./upload/report/fin.xlsx')
    sheet_select = wb['失信与被执行人']
    sheet_select['A1'] = '失信人: '
    sheet_select['A2'] = 'None'
    sheet_select['B1'] = '被执行人: '
    sheet_select['B2'] = 'None'



    return render(request, 'finace/output.html', {'name': name, 'asset': asset, 'debt': debt, 'equilt': equilt,
                                                  'risk': risk, 'return': re})

def file_download(request):
    with open('./upload/report/finance.xlsx', 'rb') as f:
        c = f.read()
    file = HttpResponse(c)
    file['Content-Type'] = 'application/vnd.ms-excel'
    file['Content-Disposition'] = 'attachment;filename = "report.xlsx"'
    return file

